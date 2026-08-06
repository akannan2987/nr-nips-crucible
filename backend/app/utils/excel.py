"""Excel/CSV reading helpers shared by the upload endpoints.

The legacy (v1) backend read spreadsheets with the `xlsx` library using
`raw:false, defval:''`, which yields **formatted strings** for every cell
(this is what prevents CAS numbers like "58-08-2" being mangled into date
serials). openpyxl returns typed Python values instead, so `format_cell`
converts them to the closest string the original backend would have produced.
"""

import csv
import io
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


def format_cell(value: Any) -> str:
    """Convert an openpyxl cell value to a display string (xlsx raw:false)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        # 194.0 → "194" (xlsx formats integral floats without the .0)
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, datetime):
        # European day-first format, matching the SLIMS exports this app
        # ingests (the samples parser normalises DD/MM/YYYY → ISO).
        if value.hour or value.minute or value.second:
            return value.strftime("%d/%m/%Y %H:%M")
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def sheet_rows_as_dicts(content: bytes) -> list[dict[str, str]]:
    """First sheet → list of row dicts keyed by the row-1 headers.

    Equivalent of `XLSX.utils.sheet_to_json(ws, {raw:false, defval:''})`.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    headers = [format_cell(h) for h in header]
    out: list[dict[str, str]] = []
    for row in rows:
        d: dict[str, str] = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            v = format_cell(row[i]) if i < len(row) else ""
            if v != "":
                empty = False
            d[h] = v
        if not empty:
            out.append(d)
    return out


def sheet_as_grid(content: bytes) -> tuple[list[list[str]], str]:
    """First sheet → (array-of-arrays of strings, sheet name).

    Equivalent of `sheet_to_json(ws, {header:1, raw:false, defval:''})`,
    used by the SLIMS samples parser which handles its own header rows.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    grid = [[format_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    return grid, ws.title


def parse_csv_rows(text: str) -> list[dict[str, str]]:
    """CSV → list of row dicts, matching the legacy manual CSV parser.

    The v1 code implemented a small RFC-4180 parser that TRIMS every field;
    Python's csv module handles the quoting, and we add the trim.
    """
    lines = [l for l in text.splitlines() if l.strip()]
    if len(lines) < 2:
        return []
    reader = csv.reader(lines)
    parsed = [[cell.strip() for cell in row] for row in reader]
    headers = parsed[0]
    out = []
    for values in parsed[1:]:
        row = {h: (values[i] if i < len(values) else "") for i, h in enumerate(headers)}
        out.append(row)
    return out
